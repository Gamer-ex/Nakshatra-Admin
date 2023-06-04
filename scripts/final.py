from openpyxl import Workbook,load_workbook
import json

wb=load_workbook('pay2.xlsx')
ws=wb.active

for i in range (2,679):
    # notes=json.loads(ws[f'T{i}'].value)
    # ws[f'L{i}'].value=notes["eventid"]
    # ws[f'M{i}'].value=notes["eventname"]
    # ws[f'N{i}'].value=notes["username"]
    # ws[f'O{i}'].value=notes["whatsapp"]
    if(ws[f'a{i}'].value==""):
        ws.delete_rows(idx=i)
        

wb.save("pay2.xlsx")