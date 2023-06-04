from openpyxl import Workbook,load_workbook

wb=load_workbook('rzp_report.xlsx')
ws=wb.active

payments=[]

for x in range (2,324):
  if ws[f'L{x}'].value==1:
    payments.append(ws[f'A{x}'].value)


print(len(payments))

# print(payments)