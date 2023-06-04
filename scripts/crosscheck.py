from openpyxl import Workbook,load_workbook
import json

wb=load_workbook('rzp.xlsx')
ws=wb.active

temp=[]
notes=[]

for i in range (2,398):
    if ws[f'L{i}'].value==1:
        temp.append(ws[f'A{i}'].value)
        a=json.loads(ws[f'U{i}'].value)
        notes.append(a["eventid"])

# for i in range (0,70):
#     ctr=0
#     if i+1>=10:
#         id=f'NK0{i+1}'
#     else:
#         id=f'NK00{i+1}'
    
#     for i in range (0,len(notes)):
#         if notes[i]==id:
#             ctr+=1

#     print(f'{id}::{ctr}')
        
id="NK018"        
    
for i in range (0,len(notes)):
    if notes[i]==id:
        print(temp[i])
        